from openpyxl import Workbook


## WING

# === Création du fichier Excel ===
wb = Workbook()

# === Création de plusieurs feuilles ===
ws1 = wb.active              # Feuille active par défaut
ws1.title = "parameters"
ws2 = wb.create_sheet(title="outputs")
ws3 = wb.create_sheet(title="chain")
ws4 = wb.create_sheet(title="settings")

####
cas = "dome"
entrees = range(0,94)#choix(cas)[0]
sorties = choix(cas)[1]

# WS1
# Remplir les colonnes avec les données d'entrée
ws1.append(["name","type","lower-bound","upper-bound","values","reference","file","row","column"])
for i in entrees:
    ws1.append([f"X{i+1}", "real", 0, 1.1, "",0.5,"input",i+1,1])

# WS2
ws2.append(["name","type","virtual-evaluation","conditioning","objective","weight","exponent","objective-group","reference","crit-lower-bound","lower-bound","lb-weight","lb-reference","lb-monitoring","lb-activated-by","upper-bound","crit-upper-bound","ub-weight","ub-reference","ub-monitoring","ub-activated-by","epsilon","file","row","column"])

ws2.append(["S_SIMULATION","conditioning-success"       ,"","","","","","","","","","","","","","","","","","","","","success",1,1])
ws2.append(["S_PARAMETRIC","parametric-success"         ,"","","","","","","","","","","","","","","","","","","","","success",2,1])
ws2.append(["mass","real","","S_SIMULATION"             ,"minimize",1,"","" ,5000,"","","","","","","","","","","","","","responses",1,1])
ws2.append(["max.stress","real","","S_SIMULATION"       ,"","","", ""       , 0.01,-1,-0.5,"","","","",0.5,1,"","","","","","responses",2,1])
ws2.append(["max.displacement","real","","S_SIMULATION" ,"","","", ""       , 0.0001,-0.002,-0.001,"","","","",0.001,0.002,"","","","","","responses",3,1])
ws2.append(["max.deflection","real","","S_SIMULATION"   ,"","","", ""       , 0.0001,-0.002,-0.001,"","","","",0.001,0.002,"","","","","","responses",4,1])

# WS3
ws3.append(["Commands","","","","","","Files","","","","Resources"])
ws3.append(["name","command-line","precedence","resources","expected-duration","","name","type","path","","name","amount","description"])
ws3.append(["structural", "python C:\\Users\\liamh\\Documents\\structural_analysis_code\\structural_design.py --example wing", "", "", 1, "", "input", "input", "simulation_input.txt", "", "", "", ""])
ws3.append(["", "", "", "", "", "", "success", "output", "simulation_success.txt", "", "", "", ""])
ws3.append(["", "", "", "", "", "", "responses", "output", "simulation_output.txt", "", "", "", ""])
# python C:\\Users\\liamh\\Documents\\structural_analysis_code\\structural_design.py --example wing
# __MINAMO_TESTCASES__ structural -i input.txt -o output.txt --example wing --element frame -s success.txt --grouping no -np 162 -nr 4 -ns 2
# WS4
ws4.append(["","", "", ""])
ws4.append(["","parameters", "parameters.A1"])
ws4.append(["","outputs", "outputs.A1"])
ws4.append(["","mission", "outputs.A1"])
ws4.append(["","chain-commands", "chain.A2"])
ws4.append(["","chain-files", "chain.G2"])
ws4.append(["","chain-resources", "chain.K2"])


# === Sauvegarder le fichier ===
wb.save(r"C:\Minamo\4.3.0-r\64\msvc\19.26.28806\python39\release\examples\test_cases\TestPlatformGeneric/domeIM.xlsx")




## DOME

# === Création du fichier Excel ===
wb = Workbook()

# === Création de plusieurs feuilles ===
ws1 = wb.active              # Feuille active par défaut
ws1.title = "parameters"
ws2 = wb.create_sheet(title="outputs")
ws3 = wb.create_sheet(title="chain")
ws4 = wb.create_sheet(title="settings")

####
cas = "dome"
entrees = choix(cas)[0]
sorties = choix(cas)[1]

# WS1
# Remplir les colonnes avec les données d'entrée
ws1.append(["name","type","lower-bound","upper-bound","values","reference","file","row","column"])
for i in entrees:
    ws1.append([f"X{i+1}", "real", 0, 1.1, "",0.5,"input",i+1,1])

# WS2
ws2.append(["name","type","virtual-evaluation","conditioning","objective","weight","exponent","objective-group","reference","crit-lower-bound","lower-bound","lb-weight","lb-reference","lb-monitoring","lb-activated-by","upper-bound","crit-upper-bound","ub-weight","ub-reference","ub-monitoring","ub-activated-by","epsilon","file","row","column"])

ws2.append(["S_SIMULATION","conditioning-success"       ,"","","","","","","","","","","","","","","","","","","","","success",1,1])
ws2.append(["S_PARAMETRIC","parametric-success"         ,"","","","","","","","","","","","","","","","","","","","","success",2,1])
ws2.append(["mass","real","","S_SIMULATION"             ,"minimize",1,"","" ,5000,"","","","","","","","","","","","","","responses",1,1])
ws2.append(["max.stress","real","","S_SIMULATION"       ,"","","", ""       , 0.01,-1,-0.5,"","","","",0.5,1,"","","","","","responses",2,1])
ws2.append(["max.displacement","real","","S_SIMULATION" ,"","","", ""       , 0.0001,-0.002,-0.001,"","","","",0.001,0.002,"","","","","","responses",3,1])
ws2.append(["max.deflection","real","","S_SIMULATION"   ,"","","", ""       , 0.0001,-0.002,-0.001,"","","","",0.001,0.002,"","","","","","responses",4,1])

# WS3
ws3.append(["Commands","","","","","","Files","","","","Resources"])
ws3.append(["name","command-line","precedence","resources","expected-duration","","name","type","path","","name","amount","description"])
ws3.append(["structural", "python C:\\Users\\liamh\\Documents\\structural_analysis_code\\structural_design.py --example dome", "", "", 1, "", "input", "input", "simulation_input.txt", "", "", "", ""])
ws3.append(["", "", "", "", "", "", "success", "output", "simulation_success.txt", "", "", "", ""])
ws3.append(["", "", "", "", "", "", "responses", "output", "simulation_output.txt", "", "", "", ""])
# python C:\\Users\\liamh\\Documents\\structural_analysis_code\\structural_design.py --example wing
# __MINAMO_TESTCASES__ structural -i input.txt -o output.txt --example wing --element frame -s success.txt --grouping no -np 162 -nr 4 -ns 2
# WS4
ws4.append(["","", "", ""])
ws4.append(["","parameters", "parameters.A1"])
ws4.append(["","outputs", "outputs.A1"])
ws4.append(["","mission", "outputs.A1"])
ws4.append(["","chain-commands", "chain.A2"])
ws4.append(["","chain-files", "chain.G2"])
ws4.append(["","chain-resources", "chain.K2"])


# === Sauvegarder le fichier ===
wb.save(r"C:\Minamo\4.3.0-r\64\msvc\19.26.28806\python39\release\examples\test_cases\TestPlatformGeneric/dome.xlsx")





## Fichier input

# Chemin de destination du fichier
file_path = r"C:\Minamo\4.3.0-r\64\msvc\19.26.28806\python39\release\examples\test_cases\TestPlatformGeneric\simulation_input.txt"

# Création et écriture du fichier avec 696 zéros en colonne
with open(file_path, "w") as file:
    for _ in range(696):
        file.write("0\n")

print(f"Fichier créé avec succès à l'emplacement : {file_path}")




